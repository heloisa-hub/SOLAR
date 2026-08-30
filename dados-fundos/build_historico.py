# -*- coding: utf-8 -*-
"""
Le TODOS os Demonstrativos Mensais (formato Singulare/QI) de cada fundo,
nao so o mes mais recente, e monta o historico completo mes a mes numa
planilha -- uma aba por fundo, com Ano%/12M%/Desde-Inicio% recalculados
por composicao geometrica a partir do retorno de cada mes individual
(mais confiavel que confiar nos campos "Ano"/"12M" impressos no PDF, que
dependem de quantos meses de historico a classe ja tinha naquele mes).

Uso:
    python build_historico.py

Le da pasta bruta "5. Relatorios Fundos/<FUNDO>/" (fora do repo git,
irmã de "4. Site/"), nao da pasta curada dados-fundos/relatorios/.
"""
import os
import re
import sys
import glob
import json
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)
from parse_relatorios import parse_singulare, clean_currency, to_float_pct  # noqa: E402

RAW_ROOT = os.path.abspath(os.path.join(BASE_DIR, "..", "..", "5. Relatórios Fundos"))
OUT_XLSX = os.path.join(BASE_DIR, "Historico-Mensal-Fundos_2026-08-29.xlsx")
OUT_JSON = os.path.join(BASE_DIR, "historico.json")

# Candidatos amplos de nome de classe, cobrindo variantes historicas que
# nao aparecem mais no manifest.json (que so guarda as classes ATIVAS).
# Ordenados do mais especifico pro mais generico -- "Subordinada Junior"
# precisa ser testado antes de "Subordinada" (senao vira falso-positivo).
CANDIDATE_CLASSES_RAW = [
    "Subordinada Junior", "Subordinada Júnior", "Subordinada Mezanino",
    "Sênior10", "Sênior9", "Sênior8", "Sênior7", "Sênior6", "Sênior5",
    "Sênior4", "Sênior3", "Sênior2", "Sênior1", "Sênior",
    "Mezanino2", "Mezanino1", "Mezanino",
    "Subordinada", "Única",
]
CANDIDATE_CLASSES = sorted(set(CANDIDATE_CLASSES_RAW), key=len, reverse=True)

FUNDS = [
    {
        "slug": "solar-fidc-multissetorial",
        "nome": "Solar FIDC Multissetorial",
        "dirs": ["SOLAR FIDC", "SOLAR MULTICEDENTE FIDC"],
        "skip_contains": ["solar-br", "relatorio-solar-br", "relatório solar br"],
    },
    {
        "slug": "solar-puglia-fidc-rl",
        "nome": "Solar Puglia FIDC",
        "dirs": ["SOLAR PUGLIA FIDC"],
        "skip_contains": [],
    },
    {
        "slug": "solar-vialoc-fidc",
        "nome": "Solar Vialoc FIDC",
        "dirs": ["SOLAR VIALOC FIDC"],
        "skip_contains": [],
    },
    {
        "slug": "solar-belmonte-fidc",
        "nome": "Solar Belmonte FIDC",
        "dirs": ["SOLAR BELMONTE FIDC"],
        "skip_contains": [],
    },
]

MONTH_RE = re.compile(r"(20\d{2})[-_]?(\d{2})(?:[^\d]|$)")


def extract_period(filename):
    """'Demonstrativo-Mensal-SOLAR-Vialoc-FIDC_2021-04.pdf' -> '2021-04'
    'Demonstrativo Mensal - SOLAR VIALOC FIDC - 202607.pdf' -> '2026-07'
    Procura o ULTIMO trecho AAAA-MM ou AAAAMM no nome do arquivo -- os
    relatorios tambem trazem outras datas (ex: geracao do PDF) que podem
    aparecer antes, entao pegar a primeira ocorrencia seria arriscado.
    """
    matches = list(MONTH_RE.finditer(filename))
    if not matches:
        return None
    y, m = matches[-1].group(1), matches[-1].group(2)
    if not (1 <= int(m) <= 12):
        return None
    return f"{y}-{m}"


def find_files(fund):
    files = []
    for d in fund["dirs"]:
        full = os.path.join(RAW_ROOT, d)
        if not os.path.isdir(full):
            continue
        for p in glob.glob(os.path.join(full, "*.pdf")):
            base = os.path.basename(p).lower()
            if "demonstrativo" not in base:
                continue
            if any(s in base for s in fund["skip_contains"]):
                continue
            period = extract_period(os.path.basename(p))
            if not period:
                print(f"  [!] Não consegui extrair período de: {os.path.basename(p)}")
                continue
            files.append((period, p))
    # um arquivo por periodo -- se houver duplicata (ex: pasta curada E
    # pasta bruta tem o mesmo mes), fica o primeiro encontrado
    seen = {}
    for period, p in sorted(files):
        if period not in seen:
            seen[period] = p
    return sorted(seen.items())


def compound(returns_pct):
    """Lista de retornos mensais em % -> retorno acumulado composto em %."""
    acc = 1.0
    for r in returns_pct:
        acc *= (1 + r / 100.0)
    return (acc - 1) * 100.0


def main():
    import pdfplumber

    workbook_data = {}  # slug -> {classe: [(periodo, mes_pct, pl_reais, cdi_mes), ...]}
    warnings_log = []

    for fund in FUNDS:
        print(f"\n=== {fund['nome']} ===")
        files = find_files(fund)
        print(f"{len(files)} arquivos de Demonstrativo Mensal encontrados")
        by_class = defaultdict(dict)  # classe -> {periodo: (mes_pct, pl_reais, percentPL, cdi_mes)}

        for period, path in files:
            try:
                with pdfplumber.open(path) as pdf:
                    text = "\n".join(page.extract_text() or "" for page in pdf.pages[:1])
            except Exception as e:
                warnings_log.append(f"{fund['nome']} {period}: erro ao abrir PDF ({e})")
                continue

            parsed = parse_singulare(text, CANDIDATE_CLASSES)
            for w in parsed["warnings"]:
                pass  # warnings esperados (classe candidata que não existe naquele mês) -- não logar, é ruído
            for cname, cdata in parsed["classes"].items():
                mes = cdata.get("mes")
                if mes is None:
                    continue
                by_class[cname][period] = (
                    mes,
                    cdata.get("plReais"),
                    cdata.get("percentPL"),
                    cdata.get("cdiMes"),
                )
            if not parsed["classes"]:
                warnings_log.append(f"{fund['nome']} {period} ({os.path.basename(path)}): nenhuma classe reconhecida")

        for cname, periods in by_class.items():
            print(f"  {cname}: {len(periods)} meses ({min(periods)} a {max(periods)})")

        workbook_data[fund["slug"]] = {"nome": fund["nome"], "by_class": dict(by_class)}

    # ---- monta planilha ----
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for fund in FUNDS:
        data = workbook_data[fund["slug"]]
        ws = wb.create_sheet(fund["nome"][:31])
        headers = ["Período", "Classe", "PL da classe", "% do PL", "Mês %", "Ano %", "12M %", "Desde o Início %", "% CDI (mês)"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for cname, periods in sorted(data["by_class"].items()):
            ordered_periods = sorted(periods.keys())
            monthly_returns = [periods[p][0] for p in ordered_periods]
            for idx, period in enumerate(ordered_periods):
                mes_pct, pl_reais, percent_pl, cdi_mes = periods[period]
                ano_atual = period[:4]
                # janela "ano": todo mes deste ano ate aqui (inclusive)
                idx_ano_inicio = next(i for i, p in enumerate(ordered_periods) if p.startswith(ano_atual))
                ano_pct = compound(monthly_returns[idx_ano_inicio:idx + 1])
                # janela 12M: ultimos 12 meses terminando aqui (só se houver 12 completos)
                if idx + 1 >= 12:
                    doze_m_pct = compound(monthly_returns[idx - 11:idx + 1])
                else:
                    doze_m_pct = None
                desde_inicio_pct = compound(monthly_returns[:idx + 1])

                ws.append([
                    period, cname, pl_reais, percent_pl,
                    round(mes_pct, 2), round(ano_pct, 2),
                    round(doze_m_pct, 2) if doze_m_pct is not None else None,
                    round(desde_inicio_pct, 2),
                    cdi_mes,
                ])
        for col, width in zip("ABCDEFGHI", [10, 18, 16, 10, 10, 10, 10, 16, 12]):
            ws.column_dimensions[col].width = width

    warn_ws = wb.create_sheet("Avisos")
    warn_ws.append(["Aviso"])
    warn_ws["A1"].font = Font(bold=True)
    for w in warnings_log:
        warn_ws.append([w])
    warn_ws.column_dimensions["A"].width = 100

    wb.save(OUT_XLSX)
    print(f"\n[OK] Planilha salva em: {OUT_XLSX}")
    print(f"{len(warnings_log)} avisos (aba 'Avisos') -- revisar antes de considerar definitivo.")

    # Mesma serie, em JSON, no formato que js/funds-data.js espera por
    # classe (historicoMensal) -- consumido por parse_relatorios.py pra
    # alimentar o grafico de rentabilidade no site (antes so o template
    # "solar-br" tinha essa serie; agora todo mundo tem, vindo daqui).
    json_out = {}
    for fund in FUNDS:
        data = workbook_data[fund["slug"]]
        json_out[fund["slug"]] = {
            cname: [{"mes": p, "rentabilidade": round(periods[p][0], 2)} for p in sorted(periods.keys())]
            for cname, periods in data["by_class"].items()
        }
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(json_out, f, ensure_ascii=False, indent=2)
    print(f"[OK] Serie mensal por classe salva em: {OUT_JSON}")


if __name__ == "__main__":
    main()
